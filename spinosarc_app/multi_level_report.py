"""
Multi-level report generation for SpinoSarc.

Produces an Excel workbook and a PDF summary from the MultiLevelAnalyzer output:
  - one row per lumbar IVD level (muscle CSA/FF, dural sac CSA, stenosis flag)
  - L3-based sarcopenia (PMI, TPA, risk)
  - stenosis flags highlighted; L5-S caveat and "approx" (gap) notes included

Style matches the existing single-slice report (navy 0A4F8C headers).

IMPORTANT: stenosis is reported as a NEUTRAL measurement flag, never as a
definitive diagnosis. The radiologist interprets it.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

# Brand colors (match existing report)
NAVY = "0A4F8C"
NAVY_RGB = (10, 79, 140)
DANGER = "C0392B"
WARNING = "E67E22"
SUCCESS = "1E8449"
GREY = "666666"
LIGHT_GREY = "CCCCCC"

# Muscle display order (so reports are consistent)
MUSCLE_ORDER = [
    "multifidus_R", "multifidus_L",
    "erector_R", "erector_L",
    "psoas_R", "psoas_L",
    "QL_R", "QL_L",
]

IVD_LEVELS = ["L1-L2", "L2-L3", "L3-L4", "L4-L5", "L5-S"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _muscle_lookup(muscles):
    """List of muscle dicts -> {name: dict}."""
    out = {}
    for m in muscles:
        out[m.get("name")] = m
    return out


def _collect_stenosis_summary(result):
    """Return a list of (level, csa, category, label, caveat) for flagged levels."""
    flagged = []
    for lvl in IVD_LEVELS:
        data = result["levels"].get(lvl)
        if not data:
            continue
        sten = data.get("stenosis")
        if sten and sten.get("flag"):
            flagged.append((
                lvl,
                data.get("canal_csa_mm2"),
                sten.get("category"),
                sten.get("label"),
                sten.get("caveat"),
            ))
    return flagged


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def export_multi_level_excel(result, demographics, patient_id, out_path):
    """Write a multi-level Excel report.

    Parameters
    ----------
    result
        Output of MultiLevelAnalyzer.analyze_all().
    demographics
        Dict-like with age/sex/height/weight (or None). Used for the header.
    patient_id
        Patient identifier string.
    out_path
        Destination .xlsx path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Level Report"

    # Styles
    title_font = Font(name="Arial", bold=True, size=14, color=NAVY)
    sub_font = Font(name="Arial", italic=True, size=10, color=GREY)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=NAVY)
    section_font = Font(name="Arial", bold=True, size=11, color=NAVY)
    bold = Font(name="Arial", bold=True)
    normal = Font(name="Arial")
    danger_font = Font(name="Arial", bold=True, color=DANGER)
    thin = Side(border_style="thin", color=LIGHT_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    # --- Title ---
    ws["A1"] = "SpinoSarc Multi-Level Quantitative Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = "Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A2"].font = sub_font
    ws.merge_cells("A2:H2")

    row = 4

    # --- Demographics ---
    ws.cell(row=row, column=1, value="Patient Demographics").font = section_font
    row += 1
    demo = demographics or {}
    def _d(k, default="-"):
        v = demo.get(k) if isinstance(demo, dict) else None
        return v if v not in (None, "", 0) else default
    demo_rows = [
        ("Patient ID", patient_id or "-"),
        ("Age", _d("age")),
        ("Sex", _d("sex")),
        ("Height (cm)", _d("height_cm")),
        ("Weight (kg)", _d("weight_kg")),
    ]
    for label, val in demo_rows:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=str(val)).font = normal
        row += 1

    row += 1

    # --- Per-level table ---
    ws.cell(row=row, column=1, value="Per-Level Measurements").font = section_font
    row += 1

    headers = ["Level", "Axial Slice", "Dural Sac CSA (mm2)",
               "Stenosis Flag", "Total Muscle CSA (mm2)",
               "Mean Fat Fraction (%)", "Notes"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center
    row += 1

    for lvl in IVD_LEVELS:
        data = result["levels"].get(lvl, {})
        ax_idx = data.get("axial_slice_idx")
        csa = data.get("canal_csa_mm2")
        sten = data.get("stenosis")
        muscles = data.get("muscles", [])

        # Total muscle CSA + mean FF
        total_csa = sum(m.get("csa_mm2", 0) for m in muscles) if muscles else None
        if muscles:
            ffs = [m.get("fat_fraction", 0) for m in muscles]
            mean_ff = (sum(ffs) / len(ffs) * 100.0) if ffs else None
        else:
            mean_ff = None

        notes = []
        if data.get("approx"):
            notes.append(f"approx: nearest slice {data.get('approx_distance_mm', 0):.1f} mm away")
        if sten and sten.get("caveat"):
            notes.append(sten["caveat"])
        if data.get("error"):
            notes.append(data["error"])

        sten_label = sten["label"] if sten else "-"

        vals = [
            lvl,
            (ax_idx + 1) if ax_idx is not None else "-",
            f"{csa:.0f}" if csa is not None else "-",
            sten_label,
            f"{total_csa:.0f}" if total_csa is not None else "-",
            f"{mean_ff:.1f}" if mean_ff is not None else "-",
            "; ".join(notes) if notes else "",
        ]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            # Highlight stenosis flag cell
            if col == 4 and sten and sten.get("flag"):
                c.font = danger_font
            else:
                c.font = normal
        row += 1

    row += 1

    # --- Detailed muscle breakdown per level ---
    ws.cell(row=row, column=1, value="Muscle Detail by Level").font = section_font
    row += 1

    mdetail_headers = ["Level", "Muscle", "CSA (mm2)", "Fat Fraction (%)"]
    for col, h in enumerate(mdetail_headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center
    row += 1

    for lvl in IVD_LEVELS:
        data = result["levels"].get(lvl, {})
        muscles = data.get("muscles", [])
        if not muscles:
            continue
        lookup = _muscle_lookup(muscles)
        for mname in MUSCLE_ORDER:
            m = lookup.get(mname)
            if not m:
                continue
            vals = [
                lvl,
                mname,
                f"{m.get('csa_mm2', 0):.0f}",
                f"{m.get('fat_fraction', 0) * 100:.1f}",
            ]
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = border
                c.font = normal
            row += 1

    row += 1

    # --- Sarcopenia ---
    ws.cell(row=row, column=1, value="Sarcopenia (L3)").font = section_font
    row += 1
    sarc = result.get("sarcopenia")
    if sarc and sarc.get("result"):
        r = sarc["result"]
        sarc_rows = [
            ("Level used", sarc.get("level_used", "-")),
            ("Axial slice", (sarc.get("axial_slice_idx") + 1)
                if sarc.get("axial_slice_idx") is not None else "-"),
            ("Total Psoas Area (cm2)", r.get("total_psoas_area_cm2", "-")),
            ("PMI (cm2/m2)", r.get("pmi_cm2_per_m2") if r.get("pmi_cm2_per_m2") else "N/A"),
            ("Risk category", r.get("risk_category", "Unknown")),
            ("Note", sarc.get("note", "")),
        ]
        for label, val in sarc_rows:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=2, value=str(val)).font = normal
            row += 1
        # Sarcopenia notes
        for note in (r.get("notes") or []):
            ws.cell(row=row, column=1, value="").font = normal
            c = ws.cell(row=row, column=2, value=note)
            c.font = Font(name="Arial", italic=True, size=9, color=GREY)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Sarcopenia not computed").font = normal
        row += 1

    row += 2

    # --- Disclaimer ---
    disclaimer = (
        "DISCLAIMER: Stenosis flags are neutral measurements against literature "
        "thresholds (absolute <75, relative <100, early <130 mm2), NOT a diagnosis. "
        "Supine MRI may underestimate canal narrowing vs axial-loaded/standing. "
        "L5-S thresholds are less reliable (few cauda equina rootlets). "
        "PMI thresholds are CT-derived; MRI may differ ~5-10%. "
        "Sarcopenia requires clinical correlation (EWGSOP2). "
        "For research use; not a substitute for radiologist interpretation."
    )
    c = ws.cell(row=row, column=1, value=disclaimer)
    c.font = Font(name="Arial", italic=True, size=9, color=GREY)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=7)

    # Column widths
    widths = [16, 12, 20, 26, 22, 20, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(str(out_path))
    return str(out_path)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def export_multi_level_pdf(result, demographics, patient_id, out_path):
    """Write a multi-level PDF summary using matplotlib.

    Single-page layout: header, per-level table, stenosis summary, sarcopenia,
    disclaimer.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    navy = tuple(c / 255.0 for c in NAVY_RGB)
    danger = (0.753, 0.224, 0.169)
    grey = (0.4, 0.4, 0.4)

    with PdfPages(str(out_path)) as pdf:
        fig = plt.figure(figsize=(8.27, 11.69))  # A4 portrait
        fig.subplots_adjust(left=0.06, right=0.94, top=0.95, bottom=0.05)
        ax = fig.add_axes([0, 0, 1, 1])
        ax.axis("off")

        y = 0.96

        # --- Title ---
        ax.text(0.06, y, "SpinoSarc Multi-Level Quantitative Report",
                fontsize=16, fontweight="bold", color=navy,
                transform=ax.transAxes)
        y -= 0.022
        ax.text(0.06, y, "Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M"),
                fontsize=8, style="italic", color=grey, transform=ax.transAxes)
        y -= 0.03

        # --- Demographics line ---
        demo = demographics or {}
        def _d(k, default="-"):
            v = demo.get(k) if isinstance(demo, dict) else None
            return v if v not in (None, "", 0) else default
        demo_str = (f"Patient: {patient_id or '-'}    "
                    f"Age: {_d('age')}    Sex: {_d('sex')}    "
                    f"Height: {_d('height_cm')} cm    Weight: {_d('weight_kg')} kg")
        ax.text(0.06, y, demo_str, fontsize=9, color="black",
                transform=ax.transAxes)
        y -= 0.03

        # --- Per-level table ---
        ax.text(0.06, y, "Per-Level Measurements", fontsize=11,
                fontweight="bold", color=navy, transform=ax.transAxes)
        y -= 0.005

        col_x = [0.06, 0.17, 0.30, 0.50, 0.70, 0.86]
        col_h = ["Level", "Slice", "Dural Sac CSA", "Stenosis", "Muscle CSA", "Mean FF"]

        # Build table data
        table_rows = []
        row_flags = []  # whether row has stenosis flag
        for lvl in IVD_LEVELS:
            data = result["levels"].get(lvl, {})
            ax_idx = data.get("axial_slice_idx")
            csa = data.get("canal_csa_mm2")
            sten = data.get("stenosis")
            muscles = data.get("muscles", [])
            total_csa = sum(m.get("csa_mm2", 0) for m in muscles) if muscles else None
            if muscles:
                ffs = [m.get("fat_fraction", 0) for m in muscles]
                mean_ff = sum(ffs) / len(ffs) * 100.0 if ffs else None
            else:
                mean_ff = None
            lvl_disp = lvl + (" *" if data.get("approx") else "")
            table_rows.append([
                lvl_disp,
                f"{ax_idx + 1}" if ax_idx is not None else "-",
                f"{csa:.0f} mm2" if csa is not None else "-",
                (sten["category"] if sten else "-"),
                f"{total_csa:.0f}" if total_csa is not None else "-",
                f"{mean_ff:.1f}%" if mean_ff is not None else "-",
            ])
            row_flags.append(bool(sten and sten.get("flag")))

        # Header row
        th = 0.022
        ax.add_patch(plt.Rectangle((0.05, y - th), 0.89, th, transform=ax.transAxes,
                                   facecolor=navy, edgecolor="none"))
        for cx, h in zip(col_x, col_h):
            ax.text(cx, y - th * 0.72, h, fontsize=8.5, fontweight="bold",
                    color="white", transform=ax.transAxes)
        y -= th

        for r_i, rdata in enumerate(table_rows):
            color = danger if row_flags[r_i] else "black"
            weight = "bold" if row_flags[r_i] else "normal"
            for cx, val in zip(col_x, rdata):
                ax.text(cx, y - th * 0.72, str(val), fontsize=8.5,
                        color=color, fontweight=weight, transform=ax.transAxes)
            # subtle row separator
            ax.plot([0.05, 0.94], [y - th, y - th], color=(0.85, 0.85, 0.85),
                    lw=0.4, transform=ax.transAxes)
            y -= th

        y -= 0.015

        # --- Stenosis summary ---
        flagged = _collect_stenosis_summary(result)
        ax.text(0.06, y, "Stenosis Summary", fontsize=11, fontweight="bold",
                color=navy, transform=ax.transAxes)
        y -= 0.022
        if flagged:
            for lvl, csa, cat, label, caveat in flagged:
                line = f"  {lvl}: {csa:.0f} mm2 - {label}"
                ax.text(0.06, y, line, fontsize=9, color=danger,
                        fontweight="bold", transform=ax.transAxes)
                y -= 0.018
                if caveat:
                    ax.text(0.09, y, caveat, fontsize=7.5, style="italic",
                            color=grey, transform=ax.transAxes)
                    y -= 0.016
        else:
            ax.text(0.06, y, "  No levels below stenosis thresholds.",
                    fontsize=9, color="black", transform=ax.transAxes)
            y -= 0.018

        y -= 0.015

        # --- Sarcopenia ---
        ax.text(0.06, y, "Sarcopenia (L3)", fontsize=11, fontweight="bold",
                color=navy, transform=ax.transAxes)
        y -= 0.022
        sarc = result.get("sarcopenia")
        if sarc and sarc.get("result"):
            r = sarc["result"]
            pmi = r.get("pmi_cm2_per_m2")
            tpa = r.get("total_psoas_area_cm2", "-")
            risk = r.get("risk_category", "Unknown")
            sline = (f"  TPA: {tpa} cm2    "
                     f"PMI: {pmi if pmi else 'N/A'} cm2/m2    Risk: {risk}")
            ax.text(0.06, y, sline, fontsize=9, color="black",
                    transform=ax.transAxes)
            y -= 0.018
            if sarc.get("note"):
                ax.text(0.06, y, "  " + sarc["note"], fontsize=7.5,
                        style="italic", color=grey, transform=ax.transAxes)
                y -= 0.016
        else:
            ax.text(0.06, y, "  Sarcopenia not computed.", fontsize=9,
                    color="black", transform=ax.transAxes)
            y -= 0.018

        if any(result["levels"].get(l, {}).get("approx") for l in IVD_LEVELS):
            y -= 0.01
            ax.text(0.06, y, "* approx: level in axial gap; nearest slice used.",
                    fontsize=7.5, style="italic", color=grey,
                    transform=ax.transAxes)
            y -= 0.016

        # --- Disclaimer (bottom) ---
        disclaimer = (
            "DISCLAIMER: Stenosis flags are neutral measurements against literature "
            "thresholds (absolute <75, relative <100, early <130 mm2), NOT a diagnosis. "
            "Supine MRI may underestimate narrowing vs standing. L5-S thresholds less "
            "reliable. PMI thresholds CT-derived; MRI may differ ~5-10%. Sarcopenia "
            "requires clinical correlation (EWGSOP2). Research use only; not a "
            "substitute for radiologist interpretation."
        )
        ax.text(0.06, 0.04, disclaimer, fontsize=7, style="italic", color=grey,
                transform=ax.transAxes, wrap=True,
                verticalalignment="bottom")

        pdf.savefig(fig)
        plt.close(fig)

    return str(out_path)
