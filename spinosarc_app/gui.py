"""SpinoSarc GUI v3 - axial + sagittal bidirectional locator + dural sac ROI (CSA)."""
import sys, os, tempfile
from pathlib import Path
import numpy as np
import nibabel as nib
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QSlider, QComboBox, QFileDialog, QFrame, QListWidget,
    QGroupBox, QFormLayout, QMessageBox, QSplitter, QTableWidget,
    QTableWidgetItem, QHeaderView, QSpinBox, QLineEdit,
    QDialog, QDialogButtonBox,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPoint
from PyQt6.QtGui import QPixmap, QImage, QColor, QPainter, QPen, QPolygon
from .analyzer import SpinoSarcAnalyzer, Demographics
from . import dicom_loader

NAVY='#0A2540'; PRIMARY='#1E5AA8'; ACCENT='#00B8D4'
SUCCESS='#0FA958'; WARNING='#F59E0B'; DANGER='#DC2626'
LIGHT='#F8FAFC'; TXT_DK='#0F172A'; TXT_MD='#475569'; TXT_LT='#94A3B8'
ROI_COLOR = (250, 204, 21)

MUSCLE_COLORS_RGB = {
    1:(239,68,68), 2:(59,130,246), 3:(16,185,129), 4:(139,92,246),
    5:(245,158,11), 6:(6,182,212), 7:(236,72,153), 8:(100,116,139),
}

def detect_orientation(nifti_path):
    img = nib.load(str(nifti_path))
    slice_dir = np.abs(img.affine[:3, 2])
    ax_idx = int(slice_dir.argmax())
    return {0:'sagittal', 1:'coronal', 2:'axial'}[ax_idx]


class ImageDisplay(QLabel):
    clicked_fraction = pyqtSignal(float, float)
    roi_changed = pyqtSignal()
    mouse_entered = pyqtSignal()

    def __init__(self, interactive_click=False):
        super().__init__()
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setMinimumSize(360, 360)
        self.setStyleSheet(f"background-color: black; border: 1px solid {TXT_LT};")
        self._img = None
        self._mask = None
        self._canal_overlay = None
        self._line_y = None
        self._line_x = None
        self._level_lines = []  # list of dicts: {y_frac, color_rgb, label, dashed}
        self._rotation = 0
        self._interactive_click = interactive_click
        self._roi_enabled = False
        self._roi_points = []
        self._roi_closed = False
        self._disp_w = 0; self._disp_h = 0
        self._pix_w = 0; self._pix_h = 0
        self._offset_x = 0; self._offset_y = 0
        self.setMouseTracking(True)

    def set_image(self, img_array, overlay_mask=None,
                  line_y_fraction=None, line_x_fraction=None, rotation=0,
                  level_lines=None, canal_overlay=None):
        self._img = img_array
        self._mask = overlay_mask
        self._canal_overlay = canal_overlay
        self._line_y = line_y_fraction
        self._line_x = line_x_fraction
        if level_lines is not None:
            self._level_lines = level_lines
        self._rotation = rotation
        self._refresh()

    def clear_level_lines(self):
        self._level_lines = []
        self._refresh()

    def set_roi_enabled(self, on):
        self._roi_enabled = on
        if not on:
            pass
        self._refresh()

    def clear_roi(self):
        self._roi_points = []
        self._roi_closed = False
        self._refresh()
        self.roi_changed.emit()

    def get_roi_points_imgcoords(self):
        if self._roi_closed and len(self._roi_points) >= 3:
            return list(self._roi_points)
        return None

    def _refresh(self):
        if self._img is None:
            return
        img = self._img
        vmin, vmax = np.percentile(img, [1, 99])
        if vmax - vmin < 1e-6:
            vmax = vmin + 1
        norm = np.clip((img - vmin)/(vmax-vmin)*255, 0, 255).astype(np.uint8)
        norm = np.rot90(norm, k=self._rotation)
        h, w = norm.shape
        self._disp_h, self._disp_w = h, w
        rgb = np.stack([norm]*3, axis=-1)

        if self._mask is not None:
            mask = np.rot90(self._mask, k=self._rotation)
            for label, color in MUSCLE_COLORS_RGB.items():
                px = (mask == label)
                if px.sum() == 0:
                    continue
                rgb[px] = (rgb[px]*0.45 + np.array(color)*0.55).astype(np.uint8)

        # Canal (dural sac) overlay - cyan, semi-transparent
        if self._canal_overlay is not None:
            try:
                canal = np.rot90(self._canal_overlay, k=self._rotation)
                if canal.shape == norm.shape:
                    canal_color = np.array([0, 200, 255])  # cyan
                    px = canal > 0
                    if px.sum() > 0:
                        rgb[px] = (rgb[px]*0.50 + canal_color*0.50).astype(np.uint8)
            except Exception:
                pass

        if self._line_y is not None:
            ly = int(np.clip(self._line_y, 0.0, 1.0) * (h - 1))
            rgb[max(0,ly-1):ly+2, :] = [220, 38, 38]
        if self._line_x is not None:
            lx = int(np.clip(self._line_x, 0.0, 1.0) * (w - 1))
            rgb[:, max(0,lx-1):lx+2] = [220, 38, 38]

        # Level lines (multi-line overlay, e.g. from TotalSpineSeg)
        # Each entry: {"y_frac": float, "color_rgb": (r,g,b), "label": str, "dashed": bool}
        for ll in self._level_lines:
            try:
                yf = float(ll.get("y_frac", 0.5))
                color = ll.get("color_rgb", (255, 255, 0))
                dashed = bool(ll.get("dashed", False))
                ly = int(np.clip(yf, 0.0, 1.0) * (h - 1))
                if dashed:
                    # Dashed: write the color every N pixels along x
                    dash_len = 8
                    gap_len = 6
                    period = dash_len + gap_len
                    xs = np.arange(w)
                    keep = (xs % period) < dash_len
                    if 0 <= ly < h:
                        rgb[ly, keep] = color
                else:
                    rgb[max(0, ly):min(h, ly + 2), :] = color
            except Exception:
                continue

        rgb = np.ascontiguousarray(rgb)
        qimg = QImage(rgb.tobytes(), w, h, w*3, QImage.Format.Format_RGB888)
        pix = QPixmap.fromImage(qimg)
        scaled = pix.scaled(self.size(), Qt.AspectRatioMode.KeepAspectRatio,
                             Qt.TransformationMode.SmoothTransformation)
        self._pix_w = scaled.width()
        self._pix_h = scaled.height()
        self._offset_x = (self.width() - self._pix_w) // 2
        self._offset_y = (self.height() - self._pix_h) // 2

        # Draw level labels on the scaled pixmap
        if self._level_lines:
            painter = QPainter(scaled)
            font = painter.font()
            font.setPointSize(10)
            font.setBold(True)
            painter.setFont(font)
            for ll in self._level_lines:
                try:
                    yf = float(ll.get("y_frac", 0.5))
                    color = ll.get("color_rgb", (255, 255, 0))
                    label_text = ll.get("label", "")
                    if not label_text:
                        continue
                    # Map y_frac to pixmap y coordinate
                    label_y = int(np.clip(yf, 0.0, 1.0) * (self._pix_h - 1))
                    # Background rectangle for readability (semi-transparent black)
                    fm = painter.fontMetrics()
                    text_w = fm.horizontalAdvance(label_text)
                    text_h = fm.height()
                    pad = 3
                    # Position label on the RIGHT edge of the pixmap
                    rect_x = self._pix_w - text_w - 2 * pad - 2
                    rect_y = label_y - text_h // 2
                    # Clamp to keep label fully visible
                    rect_y = max(0, min(self._pix_h - text_h, rect_y))
                    painter.fillRect(rect_x, rect_y, text_w + 2 * pad, text_h,
                                     QColor(0, 0, 0, 180))
                    painter.setPen(QColor(*color))
                    painter.drawText(rect_x + pad, rect_y + fm.ascent(), label_text)
                except Exception:
                    continue
            painter.end()

        if self._roi_points:
            painter = QPainter(scaled)
            pen = QPen(QColor(*ROI_COLOR))
            pen.setWidth(2)
            painter.setPen(pen)
            scr_pts = [self._imgpt_to_pixmap(c, r) for (c, r) in self._roi_points]
            for i in range(len(scr_pts) - 1):
                painter.drawLine(scr_pts[i][0], scr_pts[i][1],
                                 scr_pts[i+1][0], scr_pts[i+1][1])
            if self._roi_closed and len(scr_pts) >= 3:
                painter.drawLine(scr_pts[-1][0], scr_pts[-1][1],
                                 scr_pts[0][0], scr_pts[0][1])
            painter.setBrush(QColor(*ROI_COLOR))
            for (x, y) in scr_pts:
                painter.drawEllipse(QPoint(x, y), 3, 3)
            painter.end()

        self.setPixmap(scaled)

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._refresh()

    def _rot_dims(self):
        H0, W0 = self._img.shape
        if self._rotation % 2 == 0:
            return H0, W0
        else:
            return W0, H0

    def _imgpt_to_disp(self, col, row):
        H0, W0 = self._img.shape
        k = self._rotation % 4
        r, c = row, col
        for _ in range(k):
            r, c = (W0 - 1 - c), r
            H0, W0 = W0, H0
        return c, r

    def _disp_to_imgpt(self, x_disp, y_disp):
        H0, W0 = self._img.shape
        k = self._rotation % 4
        inv_k = (4 - k) % 4
        dh, dw = self._rot_dims()
        r, c = y_disp, x_disp
        cur_h, cur_w = dh, dw
        for _ in range(inv_k):
            r, c = (cur_w - 1 - c), r
            cur_h, cur_w = cur_w, cur_h
        return c, r

    def _imgpt_to_pixmap(self, col, row):
        xd, yd = self._imgpt_to_disp(col, row)
        if self._disp_w == 0 or self._disp_h == 0:
            return self._offset_x, self._offset_y
        sx = self._pix_w / self._disp_w
        sy = self._pix_h / self._disp_h
        return int(self._offset_x + xd * sx), int(self._offset_y + yd * sy)

    def _pixmap_to_imgpt(self, px, py):
        if self._img is None or self._pix_w == 0 or self._pix_h == 0:
            return None
        x = px - self._offset_x
        y = py - self._offset_y
        if x < 0 or y < 0 or x >= self._pix_w or y >= self._pix_h:
            return None
        xd = x * (self._disp_w / self._pix_w)
        yd = y * (self._disp_h / self._pix_h)
        col, row = self._disp_to_imgpt(xd, yd)
        return col, row

    def _pixmap_to_disp_fraction(self, px, py):
        if self._pix_w == 0 or self._pix_h == 0:
            return None
        x = px - self._offset_x
        y = py - self._offset_y
        if x < 0 or y < 0 or x >= self._pix_w or y >= self._pix_h:
            return None
        return (x / self._pix_w, y / self._pix_h)

    def mousePressEvent(self, e):
        if self._img is None:
            return
        pos = e.position()
        px, py = int(pos.x()), int(pos.y())

        if self._roi_enabled:
            if e.button() == Qt.MouseButton.LeftButton:
                if self._roi_closed:
                    self._roi_points = []
                    self._roi_closed = False
                pt = self._pixmap_to_imgpt(px, py)
                if pt is not None:
                    self._roi_points.append(pt)
                    self._refresh()
            elif e.button() == Qt.MouseButton.RightButton:
                if len(self._roi_points) >= 3:
                    self._roi_closed = True
                    self._refresh()
                    self.roi_changed.emit()
            return

        if self._interactive_click and e.button() == Qt.MouseButton.LeftButton:
            frac = self._pixmap_to_disp_fraction(px, py)
            if frac is not None:
                self.clicked_fraction.emit(frac[0], frac[1])

    def mouseDoubleClickEvent(self, e):
        if self._roi_enabled and len(self._roi_points) >= 3:
            self._roi_closed = True
            self._refresh()
            self.roi_changed.emit()

    def enterEvent(self, e):
        self.mouse_entered.emit()
        super().enterEvent(e)


class EngineLoaderThread(QThread):
    finished = pyqtSignal(object)
    error = pyqtSignal(str)
    def run(self):
        try:
            an = SpinoSarcAnalyzer(use_gpu=True)
            self.finished.emit(an)
        except Exception as e:
            self.error.emit(str(e))


class DicomLoaderThread(QThread):
    """Synapse klasörünü scan + convert et arka planda (UI takılmasın)."""
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, folder, mode='scan_and_convert',
                 work_dir=None, axial_uid=None, sagittal_uid=None):
        super().__init__()
        self.folder = folder
        self.mode = mode  # 'scan_and_convert' veya 'convert_picked'
        self.work_dir = work_dir
        self.axial_uid = axial_uid
        self.sagittal_uid = sagittal_uid

    def run(self):
        try:
            if self.mode == 'scan_and_convert':
                self.progress.emit("Scanning DICOM folder...")
                out = dicom_loader.load_synapse_set(self.folder)
            else:
                self.progress.emit("Converting selected series...")
                out = dicom_loader.convert_picked(
                    self.folder, self.work_dir, self.axial_uid, self.sagittal_uid)
            self.finished.emit(out)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.error.emit(str(e))


class PickSeriesDialog(QDialog):
    """Birden fazla axial/sagittal aday varsa kullanıcıya seçim sun."""
    def __init__(self, axial_candidates, sagittal_candidates, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select DICOM Series")
        self.setMinimumWidth(600)
        self._ax_cands = axial_candidates
        self._sag_cands = sagittal_candidates

        lay = QVBoxLayout(self)
        info = QLabel("Multiple candidates found. Please select which series to use:")
        info.setStyleSheet("font-weight: 600; color: " + NAVY + ";")
        lay.addWidget(info)

        # Axial
        ax_box = QGroupBox("Axial T2")
        ax_box.setStyleSheet("QGroupBox { font-weight: 700; color: " + PRIMARY + "; }")
        ax_lay = QVBoxLayout(ax_box)
        self.ax_combo = QComboBox()
        for c in axial_candidates:
            label = (f"#{c['series_number']}  {c['description']}  "
                     f"({c['n_slices']} slice, {c['matrix']})")
            self.ax_combo.addItem(label, userData=c['uid'])
        if not axial_candidates:
            self.ax_combo.addItem("(no candidate)", userData=None)
            self.ax_combo.setEnabled(False)
        ax_lay.addWidget(self.ax_combo)
        lay.addWidget(ax_box)

        # Sagittal
        sag_box = QGroupBox("Sagittal T2")
        sag_box.setStyleSheet("QGroupBox { font-weight: 700; color: " + PRIMARY + "; }")
        sag_lay = QVBoxLayout(sag_box)
        self.sag_combo = QComboBox()
        for c in sagittal_candidates:
            label = (f"#{c['series_number']}  {c['description']}  "
                     f"({c['n_slices']} slice, {c['matrix']})")
            self.sag_combo.addItem(label, userData=c['uid'])
        if not sagittal_candidates:
            self.sag_combo.addItem("(no candidate)", userData=None)
            self.sag_combo.setEnabled(False)
        sag_lay.addWidget(self.sag_combo)
        lay.addWidget(sag_box)

        # OK / Cancel
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def picked_uids(self):
        return self.ax_combo.currentData(), self.sag_combo.currentData()


class DropZone(QFrame):
    files_dropped = pyqtSignal(list)
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setMinimumSize(400, 250)
        self.setStyleSheet(
            "QFrame { background-color: " + LIGHT + "; "
            "border: 2px dashed " + ACCENT + "; border-radius: 12px; }"
        )
        lay = QVBoxLayout()
        icon = QLabel("File")
        icon.setStyleSheet("font-size: 36px; font-weight: 700;")
        icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title = QLabel("Drop NIfTI file(s) here")
        title.setStyleSheet("color: " + TXT_DK + "; font-size: 16px; font-weight: 600;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint = QLabel("Drop NIfTI file(s) or a Synapse DICOM folder\n"
                      "(GUI auto-detects T2 axial + sagittal)")
        hint.setStyleSheet("color: " + TXT_MD + "; font-size: 11px;")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        btn = QPushButton("Browse files / folder...")
        btn.setMaximumWidth(180)
        btn.setStyleSheet(
            "QPushButton { background-color: " + PRIMARY + "; color: white; "
            "padding: 8px 16px; border-radius: 6px; font-weight: 600; }"
            "QPushButton:hover { background-color: " + NAVY + "; }"
        )
        btn.clicked.connect(self.browse)
        lay.addStretch()
        lay.addWidget(icon)
        lay.addWidget(title)
        lay.addWidget(hint)
        lay.addSpacing(20)
        h = QHBoxLayout()
        h.addStretch()
        h.addWidget(btn)
        h.addStretch()
        lay.addLayout(h)
        lay.addStretch()
        self.setLayout(lay)
    def browse(self):
        # First check whether input is a folder or files
        from PyQt6.QtWidgets import QMessageBox
        box = QMessageBox(self)
        box.setWindowTitle("Browse")
        box.setText("What would you like to load?")
        nifti_btn = box.addButton("NIfTI file(s)", QMessageBox.ButtonRole.AcceptRole)
        folder_btn = box.addButton("DICOM folder", QMessageBox.ButtonRole.AcceptRole)
        box.addButton(QMessageBox.StandardButton.Cancel)
        box.exec()
        clicked = box.clickedButton()
        if clicked == nifti_btn:
            paths, _ = QFileDialog.getOpenFileNames(self, "Select NIfTI file(s)", "",
                                                      "NIfTI files (*.nii *.nii.gz);;All files (*)")
            if paths:
                self.files_dropped.emit(paths)
        elif clicked == folder_btn:
            folder = QFileDialog.getExistingDirectory(self, "Select DICOM folder")
            if folder:
                self.files_dropped.emit([folder])
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
    def dropEvent(self, e):
        urls = e.mimeData().urls()
        if urls:
            paths = [u.toLocalFile() for u in urls]
            self.files_dropped.emit(paths)


class SpinoSarcWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SpinoSarc")
        self.setMinimumSize(1400, 850)
        self.setStyleSheet("QMainWindow { background-color: white; }")
        self.analyzer = None
        # DICOM-mode state (her slice bagimsiz)
        self.axial_slices = []        # list of dict (per-slice DICOM data)
        # NIfTI-mode state (3D volume - dosya drop akisi icin)
        self.axial_data = None
        self.axial_affine = None
        self.axial_header = None
        self.axial_path = None
        self.sagittal_data = None
        self.sagittal_affine = None
        self.sagittal_mid_idx = None
        self.current_sag_idx = 0
        self.active_panel = 'axial'
        self.is_volume = False
        self.current_slice_idx = 0
        self.last_result = None
        self.axial_rotation = 0
        self.sagittal_rotation = 1
        # Station-based multi-station axial state
        self.axial_stations = []          # DEPRECATED - no longer used (kept for safety)
        self.current_station_idx = 0
        self._build_ui()
        self._load_engine_async()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(20, 16, 20, 16)
        main_layout.setSpacing(12)
        header = QHBoxLayout()
        title = QLabel("SpinoSarc")
        title.setStyleSheet("color: " + NAVY + "; font-size: 22px; font-weight: 700;")
        subtitle = QLabel("Paraspinal Muscle & Sarcopenia Analyzer")
        subtitle.setStyleSheet("color: " + TXT_MD + "; font-size: 12px; font-style: italic;")
        tb = QVBoxLayout(); tb.setSpacing(0)
        tb.addWidget(title); tb.addWidget(subtitle)
        # New Case butonu - tum state'i temizler, yeni hasta acmak icin
        self.new_case_btn = QPushButton("New Case")
        self.new_case_btn.setMinimumHeight(34)
        self.new_case_btn.setMinimumWidth(110)
        self.new_case_btn.setStyleSheet(
            "QPushButton { background-color: #b00020; color: white; "
            "font-weight: 600; font-size: 12px; border-radius: 6px; padding: 6px 14px; }"
            "QPushButton:hover { background-color: #d32f2f; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.new_case_btn.setEnabled(False)
        self.new_case_btn.clicked.connect(self._on_new_case)
        self.engine_status = QLabel("Loading model...")
        self.engine_status.setStyleSheet("color: " + WARNING + "; font-size: 11px; font-weight: 600;")
        header.addLayout(tb); header.addStretch()
        header.addWidget(self.new_case_btn)
        header.addSpacing(12)
        header.addWidget(self.engine_status)
        main_layout.addLayout(header)
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: " + TXT_LT + ";")
        main_layout.addWidget(sep)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.left_widget = QWidget()
        left_layout = QVBoxLayout(self.left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        self.drop_zone = DropZone()
        self.drop_zone.files_dropped.connect(self.load_files)
        left_layout.addWidget(self.drop_zone)

        self.viewer_widget = QWidget()
        viewer_layout = QVBoxLayout(self.viewer_widget)
        viewer_layout.setContentsMargins(0, 0, 0, 0)
        images_row = QHBoxLayout()
        ax_box = QVBoxLayout()
        ax_lbl = QLabel("AXIAL")
        ax_lbl.setStyleSheet("color: " + PRIMARY + "; font-weight: 700; font-size: 11px; letter-spacing: 2px;")
        ax_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ax_box.addWidget(ax_lbl)
        self.axial_display = ImageDisplay(interactive_click=False)
        self.axial_display.roi_changed.connect(self._on_roi_changed)
        self.axial_display.mouse_entered.connect(lambda: self._set_active_panel('axial'))
        ax_box.addWidget(self.axial_display)
        images_row.addLayout(ax_box, 1)
        sag_box = QVBoxLayout()
        self.sag_lbl = QLabel("SAGITTAL")
        self.sag_lbl.setStyleSheet("color: " + PRIMARY + "; font-weight: 700; font-size: 11px; letter-spacing: 2px;")
        self.sag_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sag_box.addWidget(self.sag_lbl)
        self.sagittal_display = ImageDisplay(interactive_click=True)
        self.sagittal_display.clicked_fraction.connect(self._on_sagittal_click)
        self.sagittal_display.mouse_entered.connect(lambda: self._set_active_panel('sagittal'))
        sag_box.addWidget(self.sagittal_display)
        images_row.addLayout(sag_box, 1)
        viewer_layout.addLayout(images_row)

        slider_lay = QHBoxLayout()
        self.slice_label = QLabel("Axial: -")
        self.slice_label.setStyleSheet("color: " + TXT_DK + "; font-weight: 600;")
        self.slice_slider = QSlider(Qt.Orientation.Horizontal)
        self.slice_slider.valueChanged.connect(self._on_slider)
        slider_lay.addWidget(self.slice_label)
        slider_lay.addWidget(self.slice_slider, 1)
        viewer_layout.addLayout(slider_lay)

        sag_slider_lay = QHBoxLayout()
        self.sag_slice_label = QLabel("Sagittal: -")
        self.sag_slice_label.setStyleSheet("color: " + PRIMARY + "; font-weight: 600;")
        self.sag_slider = QSlider(Qt.Orientation.Horizontal)
        self.sag_slider.setEnabled(False)
        self.sag_slider.valueChanged.connect(self._on_sag_slider)
        sag_slider_lay.addWidget(self.sag_slice_label)
        sag_slider_lay.addWidget(self.sag_slider, 1)
        viewer_layout.addLayout(sag_slider_lay)
        self.viewer_widget.hide()
        left_layout.addWidget(self.viewer_widget)

        self.file_info = QLabel("")
        self.file_info.setStyleSheet("color: " + TXT_LT + "; font-size: 11px; font-style: italic;")
        left_layout.addWidget(self.file_info)
        splitter.addWidget(self.left_widget)

        right = QWidget()
        right.setMinimumWidth(380); right.setMaximumWidth(420)
        rlay = QVBoxLayout(right)
        rlay.setContentsMargins(8, 0, 0, 0)

        demo_group = QGroupBox("Patient Demographics")
        demo_group.setStyleSheet(self._gs())
        form = QFormLayout(demo_group)
        self.patient_id_input = QLineEdit(); self.patient_id_input.setPlaceholderText("e.g. SUB-001")
        form.addRow("Patient ID:", self.patient_id_input)
        self.age_input = QSpinBox(); self.age_input.setRange(0,120); self.age_input.setSpecialValueText("-"); self.age_input.setSuffix(" yrs")
        self.sex_input = QComboBox(); self.sex_input.addItems(["-","M","F"])
        self.height_input = QSpinBox(); self.height_input.setRange(0,250); self.height_input.setSpecialValueText("-"); self.height_input.setSuffix(" cm")
        self.weight_input = QSpinBox(); self.weight_input.setRange(0,300); self.weight_input.setSpecialValueText("-"); self.weight_input.setSuffix(" kg")
        form.addRow("Age:", self.age_input); form.addRow("Sex:", self.sex_input)
        form.addRow("Height:", self.height_input); form.addRow("Weight:", self.weight_input)
        rlay.addWidget(demo_group)

        roi_group = QGroupBox("Spinal Canal CSA (auto)")
        roi_group.setStyleSheet(self._gs())
        rl = QVBoxLayout(roi_group)
        hint = QLabel("Automatically segmented from TotalSpineSeg. "
                      "Run Detect Levels, then Analyze.")
        hint.setStyleSheet("color: " + TXT_MD + "; font-size: 10px;")
        hint.setWordWrap(True)
        rl.addWidget(hint)
        self.csa_label = QLabel("Dural sac CSA: -")
        self.csa_label.setStyleSheet("color: " + NAVY + "; font-size: 15px; font-weight: 700;")
        self.csa_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        rl.addWidget(self.csa_label)
        self.stenosis_label = QLabel("")
        self.stenosis_label.setStyleSheet("font-size: 12px; font-weight: 600;")
        self.stenosis_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        rl.addWidget(self.stenosis_label)
        rlay.addWidget(roi_group)

        # ------------------------------------------------------------------
        # Lumbar Level Detection panel (TotalSpineSeg integration)
        # ------------------------------------------------------------------
        levels_group = QGroupBox("Lumbar Level Detection")
        levels_group.setStyleSheet(self._gs())
        ll = QVBoxLayout(levels_group)

        self.detect_levels_btn = QPushButton("Detect Levels")
        self.detect_levels_btn.setEnabled(False)
        self.detect_levels_btn.setMinimumHeight(36)
        self.detect_levels_btn.setStyleSheet(
            "QPushButton { background-color: " + ACCENT + "; color: white; "
            "font-size: 13px; font-weight: 700; border-radius: 6px; padding: 6px; }"
            "QPushButton:hover:enabled { background-color: " + PRIMARY + "; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.detect_levels_btn.clicked.connect(self._on_detect_levels)
        ll.addWidget(self.detect_levels_btn)

        self.levels_status_label = QLabel("Load a case to enable level detection.")
        self.levels_status_label.setStyleSheet("color: " + TXT_LT + "; font-size: 11px; font-style: italic;")
        self.levels_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.levels_status_label.setWordWrap(True)
        ll.addWidget(self.levels_status_label)

        self.levels_list = QListWidget()
        self.levels_list.setMaximumHeight(140)
        self.levels_list.itemClicked.connect(self._on_level_item_clicked)
        self.levels_list.setStyleSheet(
            "QListWidget { background-color: " + LIGHT + "; "
            "border: 1px solid " + TXT_LT + "; border-radius: 6px; font-size: 12px; }"
        )
        ll.addWidget(self.levels_list)

        rlay.addWidget(levels_group)
        # ------------------------------------------------------------------

        self.analyze_btn = QPushButton("Analyze")
        self.analyze_btn.setEnabled(False); self.analyze_btn.setMinimumHeight(48)
        self.analyze_btn.setStyleSheet(
            "QPushButton { background-color: " + PRIMARY + "; color: white; "
            "font-size: 15px; font-weight: 700; border-radius: 8px; }"
            "QPushButton:hover:enabled { background-color: " + NAVY + "; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.analyze_btn.clicked.connect(self._on_analyze)
        rlay.addWidget(self.analyze_btn)

        self.analyze_all_btn = QPushButton("Analyze All Levels")
        self.analyze_all_btn.setEnabled(False)
        self.analyze_all_btn.setMinimumHeight(40)
        self.analyze_all_btn.setStyleSheet(
            "QPushButton { background-color: " + ACCENT + "; color: white; "
            "border: none; border-radius: 8px; font-size: 14px; font-weight: 600; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.analyze_all_btn.setToolTip(
            "Run muscle + dural sac analysis at every detected lumbar level. "
            "Requires Detect Levels first."
        )
        self.analyze_all_btn.clicked.connect(self._on_analyze_all_levels)
        rlay.addWidget(self.analyze_all_btn)

        self.results_group = QGroupBox("Analysis Results")
        self.results_group.setStyleSheet(self._gs())
        res = QVBoxLayout(self.results_group)
        self.risk_label = QLabel("-")
        self.risk_label.setStyleSheet("font-size: 18px; font-weight: 700; color: gray;")
        self.risk_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        res.addWidget(self.risk_label)
        self.pmi_label = QLabel("PMI: -")
        self.pmi_label.setStyleSheet("color: " + TXT_DK + "; font-size: 13px;")
        self.pmi_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        res.addWidget(self.pmi_label)
        self.muscle_table = QTableWidget(0, 3)
        self.muscle_table.setHorizontalHeaderLabels(["Muscle", "CSA mm2", "FF %"])
        self.muscle_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.muscle_table.verticalHeader().setVisible(False)
        self.muscle_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        res.addWidget(self.muscle_table)
        self.save_pdf_btn = QPushButton("Save PDF Report")
        self.save_pdf_btn.setEnabled(False); self.save_pdf_btn.setMinimumHeight(40)
        self.save_pdf_btn.setStyleSheet(
            "QPushButton { background-color: " + ACCENT + "; color: white; "
            "font-weight: 600; border-radius: 6px; }"
            "QPushButton:hover:enabled { background-color: " + PRIMARY + "; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.save_pdf_btn.clicked.connect(self._on_save_pdf)
        self.export_excel_btn = QPushButton("Export Excel")
        self.export_excel_btn.setEnabled(False); self.export_excel_btn.setMinimumHeight(40)
        self.export_excel_btn.setStyleSheet(
            "QPushButton { background-color: #1d7a3f; color: white; "
            "font-weight: 600; border-radius: 6px; }"
            "QPushButton:hover:enabled { background-color: #145e30; }"
            "QPushButton:disabled { background-color: " + TXT_LT + "; }"
        )
        self.export_excel_btn.clicked.connect(self._on_export_excel)
        out_btn_row = QHBoxLayout()
        out_btn_row.addWidget(self.save_pdf_btn, 1)
        out_btn_row.addWidget(self.export_excel_btn, 1)
        res.addLayout(out_btn_row)
        rlay.addWidget(self.results_group)
        rlay.addStretch()
        splitter.addWidget(right)
        splitter.setSizes([950, 400])
        main_layout.addWidget(splitter, 1)

        self.status_label = QLabel("Ready - drop NIfTI file(s) to begin")
        self.status_label.setStyleSheet("color: " + TXT_MD + "; font-size: 11px; padding: 4px;")
        main_layout.addWidget(self.status_label)

    def _gs(self):
        return ("QGroupBox { font-weight: 700; color: " + PRIMARY + "; "
                "border: 1px solid " + TXT_LT + "; border-radius: 8px; "
                "margin-top: 12px; padding-top: 14px; }"
                "QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }")

    def _load_engine_async(self):
        self.loader = EngineLoaderThread()
        self.loader.finished.connect(self._on_engine_ready)
        self.loader.error.connect(self._on_engine_error)
        self.loader.start()

    def _on_engine_ready(self, an):
        self.analyzer = an
        self.engine_status.setText("Model ready (" + str(an.engine.device) + ")")
        self.engine_status.setStyleSheet("color: " + SUCCESS + "; font-size: 11px; font-weight: 600;")
        if self.axial_data is not None or self.axial_slices:
            self.analyze_btn.setEnabled(True)
            self.detect_levels_btn.setEnabled(True)
            self.new_case_btn.setEnabled(True)

    def _on_engine_error(self, err):
        self.engine_status.setText("Model failed")
        self.engine_status.setStyleSheet("color: " + DANGER + "; font-size: 11px; font-weight: 600;")
        QMessageBox.critical(self, "Model Error", "Failed to load model:\n" + err)

    def load_files(self, paths):
        """paths: list. NIfTI dosya yolları VEYA tek klasör yolu olabilir."""
        # Check if input is a folder
        if len(paths) == 1 and Path(paths[0]).is_dir():
            self._load_dicom_folder(paths[0])
            return
        # NIfTI files workflow (legacy)
        self._load_nifti_files(paths)

    def _load_dicom_folder(self, folder):
        """DICOM klasörü -> dicom_loader thread -> DICOM slices -> GUI yükle."""
        # reset previous state
        self.axial_slices = []
        self.axial_data = None
        self.axial_affine = None
        self.axial_header = None

        self._dicom_folder = folder
        self.status_label.setText("Scanning DICOM folder, please wait...")
        self.drop_zone.setEnabled(False)
        QApplication.processEvents()

        # scan + convert in background thread
        self._dicom_thread = DicomLoaderThread(folder, mode='scan_and_convert')
        self._dicom_thread.progress.connect(
            lambda msg: self.status_label.setText(msg))
        self._dicom_thread.finished.connect(self._on_dicom_loaded)
        self._dicom_thread.error.connect(self._on_dicom_error)
        self._dicom_thread.start()

    def _on_dicom_loaded(self, out):
        """dicom_loader sonucu geldi: tek aday otomatik, çoklu aday dialog."""
        self.drop_zone.setEnabled(True)
        errors = out.get('errors', [])
        ax_cands = out.get('axial_candidates', [])
        sag_cands = out.get('sagittal_candidates', [])

        # multi-candidate check
        need_dialog = (len(ax_cands) > 1) or (len(sag_cands) > 1)

        if need_dialog:
            dlg = PickSeriesDialog(ax_cands, sag_cands, self)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                self.status_label.setText("DICOM load cancelled.")
                return
            ax_uid, sag_uid = dlg.picked_uids()
            self.status_label.setText("Converting selected series...")
            QApplication.processEvents()
            self._dicom_thread2 = DicomLoaderThread(
                self._dicom_folder, mode='convert_picked',
                work_dir=out['info'].get('work_dir'),
                axial_uid=ax_uid, sagittal_uid=sag_uid)
            self._dicom_thread2.progress.connect(
                lambda msg: self.status_label.setText(msg))
            self._dicom_thread2.finished.connect(self._on_dicom_converted)
            self._dicom_thread2.error.connect(self._on_dicom_error)
            self._dicom_thread2.start()
            return

        # Single candidate case - files already converted
        self._on_dicom_converted(out)

    def _on_dicom_converted(self, out):
        slices = out.get('axial_slices', [])
        sag_nifti = out.get('sagittal_nifti')
        errors = out.get('errors', [])
        if not slices:
            QMessageBox.critical(
                self, "DICOM load failed",
                "Could not load axial T2 from DICOM folder.\n\n" + "\n".join(errors))
            self.status_label.setText("DICOM load failed.")
            return

        # DICOM-mode: per-slice listesini state'e koy
        self.axial_slices = slices
        # clear NIfTI-mode state to avoid regression
        self.axial_data = None
        self.axial_affine = None
        self.axial_header = None
        self.axial_path = None

        # Load sagittal as NIfTI (legacy flow)
        if sag_nifti is not None:
            try:
                self.sagittal_nifti_path = str(sag_nifti)
                sag_img = nib.load(str(sag_nifti))
                self.sagittal_data = sag_img.get_fdata()
                self.sagittal_affine = sag_img.affine
                n_sag = self.sagittal_data.shape[2]
                self.sagittal_mid_idx = n_sag // 2
                self.current_sag_idx = self.sagittal_mid_idx
                self.sag_slider.setEnabled(True)
                self.sag_slider.blockSignals(True)
                self.sag_slider.setRange(0, n_sag - 1)
                self.sag_slider.setValue(self.sagittal_mid_idx)
                self.sag_slider.blockSignals(False)
                self.sag_slice_label.setText(
                    f"Sagittal: {self.sagittal_mid_idx + 1} / {n_sag}")
            except Exception as e:
                import traceback; traceback.print_exc()
                self.sagittal_data = None
                self.sagittal_affine = None
        else:
            self.sagittal_data = None
            self.sagittal_nifti_path = None
            self.sagittal_affine = None
            self.sag_slider.setEnabled(False)
            self.sag_slider.setRange(0, 0)
            self.sag_slice_label.setText("Sagittal: -")

        # Axial slider - length matches DICOM slice list
        self.is_volume = True
        n = len(slices)
        self.slice_slider.setEnabled(True)
        self.slice_slider.blockSignals(True)
        self.slice_slider.setRange(0, n - 1)
        mid = n // 2
        self.current_slice_idx = mid
        self.slice_slider.setValue(mid)
        self.slice_slider.blockSignals(False)
        self.slice_label.setText(f"Axial: {mid + 1} / {n}")

        # show GUI
        self.drop_zone.hide()
        self.viewer_widget.show()
        pass  # manual ROI removed - canal is auto-segmented
        self._reset_csa_display()

        # initial render
        self.active_panel = 'axial'
        self._update_locators()

        # Info line
        first_slice = slices[0]
        info = (f"DICOM series: {n} slices  "
                f"shape={first_slice['rows']}x{first_slice['cols']}  "
                f"pixel {first_slice['pixel_spacing'][0]:.2f}x{first_slice['pixel_spacing'][1]:.2f} mm")
        if sag_nifti is not None:
            info += f"   +sagittal: {Path(str(sag_nifti)).name}"
        self.file_info.setText(info)
        self.status_label.setText(
            f"Loaded {n} DICOM slices (sorted by InstanceNumber). "
            f"Draw ROI, click Analyze.")
        self.new_case_btn.setEnabled(True)
        self.detect_levels_btn.setEnabled(True)
        if self.analyzer is not None:
            self.analyze_btn.setEnabled(True)

    def _on_dicom_error(self, err):
        self.drop_zone.setEnabled(True)
        self.status_label.setText("DICOM error: " + err[:200])
        QMessageBox.critical(self, "DICOM error",
            "DICOM loading failed:\n" + err +
            "\n\nMake sure dcm2niix is installed:\n"
            "conda install -c conda-forge dcm2niix")

    def _load_nifti_files(self, paths):
        try:
            # NIfTI dosya drop yolu: DICOM-mode state'ini temizle
            # (even if not needed, prevent mode confusion)
            self.axial_slices = []

            axial_path = None; sagittal_path = None
            for p in paths:
                orient = detect_orientation(p)
                if orient == 'axial':
                    axial_path = p
                elif orient == 'sagittal':
                    sagittal_path = p
            if axial_path is None:
                QMessageBox.warning(self, "Missing axial",
                    "An axial scan is required for analysis.\nDrop at least the axial NIfTI file.")
                return
            ax_img = nib.load(axial_path)
            ax_arr = ax_img.get_fdata()
            self.axial_data = ax_arr
            self.axial_affine = ax_img.affine
            self.axial_header = ax_img.header
            self.axial_path = axial_path

            if ax_arr.ndim == 2 or (ax_arr.ndim == 3 and ax_arr.shape[2] == 1):
                self.is_volume = False
                self.slice_slider.setEnabled(False)
                self.slice_slider.setRange(0, 0)
                self.slice_label.setText("Single slice (no slider)")
                disp = ax_arr[:, :, 0] if ax_arr.ndim == 3 else ax_arr
                self._axial_2d = disp
                self.axial_display.set_image(disp, rotation=self.axial_rotation)
            else:
                self.is_volume = True
                n = ax_arr.shape[2]
                self.slice_slider.setEnabled(True)
                self.slice_slider.setRange(0, n - 1)
                mid = n // 2
                self.current_slice_idx = mid
                self.slice_slider.setValue(mid)
                self.axial_display.set_image(ax_arr[:, :, mid], rotation=self.axial_rotation)

            if sagittal_path is not None:
                self.sagittal_nifti_path = sagittal_path
                sag_img = nib.load(sagittal_path)
                sag_arr = sag_img.get_fdata()
                self.sagittal_data = sag_arr
                self.sagittal_affine = sag_img.affine
                n_sag = sag_arr.shape[2]
                self.sagittal_mid_idx = n_sag // 2
                self.current_sag_idx = self.sagittal_mid_idx
                self.sag_slider.setEnabled(True)
                self.sag_slider.setRange(0, n_sag - 1)
                self.sag_slider.blockSignals(True)
                self.sag_slider.setValue(self.sagittal_mid_idx)
                self.sag_slider.blockSignals(False)
                self.sag_slice_label.setText("Sagittal: " + str(self.sagittal_mid_idx + 1) + " / " + str(n_sag))
                self.active_panel = 'axial'
                self._update_locators()
                self.sag_lbl.setText("SAGITTAL")
            else:
                self.sagittal_data = None; self.sagittal_affine = None
                self.sag_slider.setEnabled(False)
                self.sag_slider.setRange(0, 0)
                self.sag_slice_label.setText("Sagittal: -")
                self.sagittal_display.clear()
                self.sagittal_display.setText("No sagittal\n(drop sagittal NIfTI\nfor locator view)")
                self.sagittal_display.setStyleSheet(
                    "background-color: " + LIGHT + "; color: " + TXT_LT + "; "
                    "border: 1px dashed " + TXT_LT + "; font-style: italic;")
                self.sag_lbl.setText("SAGITTAL")

            self.drop_zone.hide(); self.viewer_widget.show()
            self._reset_csa_display()

            zooms = ax_img.header.get_zooms()
            info = (Path(axial_path).name + "  shape " + str(ax_arr.shape) +
                    "  pixel " + str(round(zooms[0], 2)) + "x" + str(round(zooms[1], 2)) + " mm")
            if sagittal_path:
                info += "   +sagittal: " + Path(sagittal_path).name
            self.file_info.setText(info)
            self.status_label.setText("File(s) loaded. Draw dural sac ROI, enter demographics, click Analyze.")
            self.new_case_btn.setEnabled(True)
            self.detect_levels_btn.setEnabled(True)
            if self.analyzer is not None:
                self.analyze_btn.setEnabled(True)
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Load error", "Failed to load files:\n" + str(e))

    def _set_active_panel(self, panel):
        if self.active_panel == panel:
            return
        self.active_panel = panel
        self._update_locators()

    def _get_current_axial_image(self):
        """Mevcut axial slice'in pixel array'ini dondur. Hem DICOM hem NIfTI mode'u
        destekler."""
        if self.axial_slices:
            # DICOM mode
            return self.axial_slices[self.current_slice_idx]['pixel_array']
        elif self.axial_data is not None:
            # NIfTI mode
            if self.is_volume:
                return self.axial_data[:, :, self.current_slice_idx]
            else:
                return self.axial_data[:, :, 0] if self.axial_data.ndim == 3 else self.axial_data
        return None

    def _get_axial_z_world(self, axial_slice_idx):
        """Verilen axial slice'in world z koordinati. DICOM mode: ImagePositionPatient.
        NIfTI mode: affine'den hesapla."""
        if self.axial_slices:
            return float(self.axial_slices[axial_slice_idx]['z_world_mm'])
        elif self.axial_affine is not None:
            return float(self.axial_affine[2, 3] + self.axial_affine[2, 2] * axial_slice_idx)
        return None

    def _get_current_pixel_spacing(self):
        """Mevcut axial slice'in (row_mm, col_mm) pixel spacing'i."""
        if self.axial_slices:
            ps = self.axial_slices[self.current_slice_idx]['pixel_spacing']
            return float(ps[0]), float(ps[1])
        elif self.axial_header is not None:
            zooms = self.axial_header.get_zooms()
            return float(zooms[0]), float(zooms[1])
        return 1.0, 1.0

    def _update_locators(self):
        # ne axial_slices ne axial_data varsa hicbir sey yok
        if not self.axial_slices and self.axial_data is None:
            return
        has_sag = self.sagittal_data is not None

        if self.active_panel == 'axial':
            ax_line_x = None
            sag_line_y = self._compute_sagittal_line_y(self.current_slice_idx) if has_sag else None
        else:
            ax_line_x = self._compute_axial_line_x(self.current_sag_idx) if has_sag else None
            sag_line_y = None

        ax_img = self._get_current_axial_image()
        if ax_img is None:
            return
        self.axial_display.set_image(ax_img, line_x_fraction=ax_line_x, rotation=self.axial_rotation)

        if has_sag:
            self.sagittal_display.set_image(
                self.sagittal_data[:, :, self.current_sag_idx],
                line_y_fraction=sag_line_y, rotation=self.sagittal_rotation)

    def _refresh_axial(self, line_x=None):
        if not self.axial_slices and self.axial_data is None:
            return
        img = self._get_current_axial_image()
        if img is None:
            return
        self.axial_display.set_image(img, line_x_fraction=line_x, rotation=self.axial_rotation)

    def _compute_axial_line_x(self, sag_slice_idx):
        """Sagittal slice -> axial'de dikey locator x-fraction.

        DICOM mode'da bu yaklaşık olur cunku per-slice affine farkli olabilir.
        Mevcut axial slice'in image_position+image_orientation'ini kullaniyoruz."""
        if self.sagittal_data is None or self.sagittal_affine is None:
            return None
        try:
            sag_aff = self.sagittal_affine
            lr_step_vec = sag_aff[:3, 2]
            lr_origin_vec = sag_aff[:3, 3]
            world_pt = lr_origin_vec + lr_step_vec * sag_slice_idx
            lr_world_mm = world_pt[0]  # L-R world coord

            if self.axial_slices:
                # DICOM mode: mevcut slice'in IPP + IOP'sini kullan
                slc = self.axial_slices[self.current_slice_idx]
                ipp = slc['image_position']
                iop = slc['image_orientation']
                ps = slc['pixel_spacing']
                cols = slc['cols']
                # Row cosine = ilk 3 IOP. Bu axial'de L-R yonu (genelde)
                row_cos = np.array(iop[:3])
                # L-R fraction: world point'in row eksenine projeksiyon
                # axial slice'da soldan saga = col yonu - ama IOP'a gore degisir
                # Burada column cosine'in en buyuk x bileseni varsa o L-R demek
                col_cos = np.array(iop[3:6])
                # Hangi eksen L-R'a yakin? row mu col mu?
                # x bileseni en buyuk olan = L-R
                if abs(row_cos[0]) > abs(col_cos[0]):
                    # row L-R yonu
                    lr_step_mm = ps[1] * row_cos[0]
                    lr_origin_x = ipp[0]
                    n_along = cols
                else:
                    # col L-R yonu (axial genelde row L-R, col A-P)
                    lr_step_mm = ps[0] * col_cos[0]
                    lr_origin_x = ipp[0]
                    n_along = slc['rows']
                if abs(lr_step_mm) < 1e-9:
                    return 0.5
                pixel_idx = (lr_world_mm - lr_origin_x) / lr_step_mm
                fraction = pixel_idx / (n_along - 1)
                return float(np.clip(fraction, 0.0, 1.0))
            else:
                # NIfTI mode (eski)
                ax_aff = self.axial_affine
                lr_axis = int(np.abs(ax_aff[0, :3]).argmax())
                lr_ax_step = ax_aff[0, lr_axis]
                lr_ax_origin = ax_aff[0, 3]
                pixel_idx = (lr_world_mm - lr_ax_origin) / lr_ax_step
                n_along = self.axial_data.shape[lr_axis]
                fraction = pixel_idx / (n_along - 1)
                return float(np.clip(fraction, 0.0, 1.0))
        except Exception:
            import traceback; traceback.print_exc()
            return 0.5

    def _compute_sagittal_line_y(self, axial_slice_idx):
        """Axial slice -> sagittal'de yatay locator y-fraction.

        Her zaman world z koordinatini kullanir - DICOM mode'da ImagePositionPatient,
        NIfTI mode'da affine."""
        if self.sagittal_data is None or self.sagittal_affine is None:
            return None
        try:
            z_world_mm = self._get_axial_z_world(axial_slice_idx)
            if z_world_mm is None:
                return None
            sag_aff = self.sagittal_affine
            si_axis = int(np.abs(sag_aff[2, :3]).argmax())
            si_step = sag_aff[2, si_axis]
            si_origin = sag_aff[2, 3]
            pixel_idx = (z_world_mm - si_origin) / si_step
            n_along = self.sagittal_data.shape[si_axis]
            raw_fraction = pixel_idx / (n_along - 1)
            # Invert to be consistent with _sagittal_yfrac_to_axial_slice:
            # screen y=0 is TOP (cranial), so cranial z must give fraction=0.
            fraction = 1.0 - raw_fraction
            return float(np.clip(fraction, 0.0, 1.0))
        except Exception:
            return 0.5

    def _sagittal_yfrac_to_axial_slice(self, y_frac):
        """Sagittal'de yatay tiklama -> en yakin axial slice index.

        DICOM mode'da: her slice'in z'sine bak, en yakini bul.
        NIfTI mode'da: affine ile hesapla."""
        if self.sagittal_data is None:
            return None
        try:
            sag_aff = self.sagittal_affine
            si_axis = int(np.abs(sag_aff[2, :3]).argmax())
            si_step = sag_aff[2, si_axis]
            si_origin = sag_aff[2, 3]
            n_along = self.sagittal_data.shape[si_axis]
            # Screen y=0 is the TOP of the display = anatomical CRANIAL = high z.
            # Empirically verified via debug logging:
            # - y_frac=0.07 (top click)    -> needs voxel ~356 -> z=+167 (cranial OK)
            # - y_frac=0.78 (bottom click) -> needs voxel  ~84 -> z=-45  (caudal OK)
            # So y_frac must be inverted before mapping to voxel index.
            pixel_idx = (1.0 - y_frac) * (n_along - 1)
            z_world_mm = si_origin + si_step * pixel_idx

            if self.axial_slices:
                # DICOM mode: en yakin z'ye sahip slice'i bul
                best_idx = 0
                best_dist = float('inf')
                for i, slc in enumerate(self.axial_slices):
                    d = abs(slc['z_world_mm'] - z_world_mm)
                    if d < best_dist:
                        best_dist = d
                        best_idx = i
                return best_idx
            elif self.axial_data is not None and self.axial_affine is not None:
                # NIfTI mode (eski)
                ax_origin = self.axial_affine[2, 3]
                ax_z_step = self.axial_affine[2, 2]
                ax_idx = (z_world_mm - ax_origin) / ax_z_step
                ax_idx = int(round(ax_idx))
                n_ax = self.axial_data.shape[2]
                return int(np.clip(ax_idx, 0, n_ax - 1))
            return None
        except Exception:
            import traceback; traceback.print_exc()
            return None

    def _on_sag_slider(self, value):
        if self.sagittal_data is None:
            return
        self.current_sag_idx = value
        n_sag = self.sagittal_data.shape[2]
        self.sag_slice_label.setText("Sagittal: " + str(value + 1) + " / " + str(n_sag))
        self.active_panel = 'sagittal'
        self._update_locators()

    def _on_sagittal_click(self, x_frac, y_frac):
        if self.sagittal_data is None or not self.is_volume:
            return
        ax_idx = self._sagittal_yfrac_to_axial_slice(y_frac)
        if ax_idx is not None and ax_idx != self.current_slice_idx:
            self.slice_slider.setValue(ax_idx)

    def _compute_canal_overlay_for_axial_slice(self, slice_idx):
        """Return a 2D boolean mask of the canal at the given axial slice,
        resampled to the axial DICOM grid using world-space coordinates."""
        canal_path = getattr(self, "canal_nifti_path", None)
        if not canal_path or not Path(canal_path).is_file():
            return None
        if not self.axial_slices:
            return None
        if slice_idx < 0 or slice_idx >= len(self.axial_slices):
            return None
        try:
            from .totalspineseg.canal_csa import resample_canal_to_axial_slice
            import numpy as _np
            mask = resample_canal_to_axial_slice(
                canal_path,
                self.axial_slices[slice_idx],
                threshold=0.5,
            )
            return mask
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"[CANAL OVERLAY] Error: {e}")
            return None

    def _on_slider(self, value):
        # iki mode'u da destekle
        if self.axial_slices:
            n = len(self.axial_slices)
        elif self.axial_data is not None and self.is_volume:
            n = self.axial_data.shape[2]
        else:
            return
        self.current_slice_idx = value
        self.slice_label.setText(f"Axial: {value + 1} / {n}")
        self.axial_display.clear_roi()
        self._reset_csa_display()
        self.active_panel = 'axial'
        self._update_locators()
        # Re-render axial with canal overlay (if Detect Levels has been run)
        self._refresh_axial_with_canal_overlay()

    def _refresh_axial_with_canal_overlay(self):
        """Re-render the current axial slice, adding the canal mask overlay
        if a canal NIfTI has been loaded by Detect Levels."""
        if not self.axial_slices:
            return
        idx = self.current_slice_idx
        if idx < 0 or idx >= len(self.axial_slices):
            return
        canal_mask = self._compute_canal_overlay_for_axial_slice(idx)
        if canal_mask is None:
            return  # no canal data, leave existing display alone
        # Re-display current axial slice with canal overlay
        try:
            img = self.axial_slices[idx]["pixel_array"]
            self.axial_display.set_image(
                img,
                rotation=self.axial_rotation,
                canal_overlay=canal_mask,
            )
        except Exception as e:
            print(f"[CANAL OVERLAY] re-render failed: {e}")

    def _on_roi_draw_toggled(self, checked):
        self.axial_display.set_roi_enabled(checked)
        if checked:
            self.status_label.setText("ROI mode ON - left-click to add points, right-click to close.")
        else:
            self.status_label.setText("ROI mode OFF.")

    def _on_roi_clear(self):
        self.axial_display.clear_roi()
        self._reset_csa_display()

    def _reset_csa_display(self):
        self.csa_label.setText("Dural sac CSA: -")
        self.stenosis_label.setText("")

    def _on_roi_changed(self):
        pts = self.axial_display.get_roi_points_imgcoords()
        if pts is None:
            self._reset_csa_display()
            return
        area_px = self._polygon_area_px(pts)
        sp_x, sp_y = self._get_current_pixel_spacing()
        csa_mm2 = area_px * sp_x * sp_y
        self.csa_label.setText("Dural sac CSA: " + str(int(round(csa_mm2))) + " mm2")
        if csa_mm2 < 75:
            self.stenosis_label.setText("SEVERE stenosis (<75 mm2)")
            self.stenosis_label.setStyleSheet("font-size: 12px; font-weight: 600; color: " + DANGER + ";")
        elif csa_mm2 < 100:
            self.stenosis_label.setText("Relative stenosis (<100 mm2)")
            self.stenosis_label.setStyleSheet("font-size: 12px; font-weight: 600; color: " + WARNING + ";")
        else:
            self.stenosis_label.setText("No stenosis (>=100 mm2)")
            self.stenosis_label.setStyleSheet("font-size: 12px; font-weight: 600; color: " + SUCCESS + ";")

    @staticmethod
    def _polygon_area_px(points):
        n = len(points)
        if n < 3:
            return 0.0
        area = 0.0
        for i in range(n):
            x1, y1 = points[i]
            x2, y2 = points[(i + 1) % n]
            area += x1 * y2 - x2 * y1
        return abs(area) / 2.0

    def _get_demographics(self):
        age = self.age_input.value() or None
        st = self.sex_input.currentText()
        sex = st if st in ("M", "F") else None
        h = self.height_input.value() or None
        w = self.weight_input.value() or None
        if any([age, sex, h, w]):
            return Demographics(age=age, sex=sex, height_cm=h, weight_kg=w)
        return None

    def _on_level_item_clicked(self, item):
        """Jump axial slider to the level's axial slice.

        Items for OUT-OF-RANGE levels carry no UserRole data, so clicking
        them is a no-op (the level is not covered by the axial volume).
        """
        ax_idx = item.data(Qt.ItemDataRole.UserRole)
        if ax_idx is None:
            return
        try:
            ax_idx = int(ax_idx)
        except (TypeError, ValueError):
            return
        # Slider triggers _on_slider, which updates current_slice_idx and refreshes display.
        self.slice_slider.setValue(ax_idx)

    def _on_detect_levels(self):
        """Run TotalSpineSeg on the current sagittal NIfTI, parse levels,
        map to axial slices, and populate the levels list.

        WARNING: This runs synchronously - the GUI will freeze for ~60 seconds
        while TotalSpineSeg performs inference. A future version will move
        this to a QThread worker.
        """
        from PyQt6.QtWidgets import QMessageBox, QListWidgetItem
        from PyQt6.QtGui import QColor

        # ---- Sanity checks ----
        if not self.axial_slices:
            QMessageBox.warning(
                self, "Detect Levels",
                "Lumbar level detection requires a DICOM case to be loaded.",
            )
            return

        # We need a sagittal NIfTI on disk. SpinoSarc already produces one
        # at load time via dcm2niix; that path is stored in self.sagittal_path
        # (NIfTI mode) or self.sagittal_nifti_for_tss (set by the DICOM loader).
        sag_nifti = getattr(self, 'sagittal_nifti_path', None)
        if not sag_nifti or not Path(sag_nifti).is_file():
            QMessageBox.critical(
                self, "Detect Levels",
                "Could not locate a sagittal NIfTI file for this case.\n"
                "Make sure both axial and sagittal series are loaded.",
            )
            return

        # ---- Lazy import (avoid GUI import overhead on startup) ----
        from .totalspineseg.runner import TotalSpineSegRunner
        from .totalspineseg.level_mapper import LevelMapper

        runner = TotalSpineSegRunner()
        if not runner.is_available():
            QMessageBox.critical(
                self, "Detect Levels",
                "TotalSpineSeg is not available.\n\n"
                "Make sure the `totalspineseg` conda environment is installed.",
            )
            return

        # ---- Run (blocking) ----
        self.detect_levels_btn.setEnabled(False)
        self.detect_levels_btn.setText("Detecting...")
        self.levels_status_label.setText(
            "Running TotalSpineSeg (this takes ~60 seconds on Apple Silicon)..."
        )
        self.levels_list.clear()
        QApplication.processEvents()

        out_dir = str(Path(tempfile.gettempdir()) / "spinosarc_tss_output")
        # Clean stale output from prior runs (avoids picking the wrong canal
        # NIfTI when a different sagittal series was processed earlier).
        try:
            import shutil
            if Path(out_dir).exists():
                shutil.rmtree(out_dir)
        except Exception as _e:
            print(f"[TSS] Could not clean out_dir: {_e}")
        try:
            print(f"[TSS] Running on {sag_nifti}")
            result = runner.run(
                sagittal_nifti_path=sag_nifti,
                output_dir=out_dir,
                device="mps",
                step1_only=True,
                iso=True,
                progress_callback=lambda m: print(f"[TSS] {m}"),
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            self.detect_levels_btn.setEnabled(True)
            self.detect_levels_btn.setText("Detect Levels")
            self.levels_status_label.setText(f"Error: {e}")
            return

        if not result["success"]:
            self.detect_levels_btn.setEnabled(True)
            self.detect_levels_btn.setText("Detect Levels")
            self.levels_status_label.setText(
                f"TotalSpineSeg failed: {result.get('error', 'unknown error')}"
            )
            print("[TSS stderr]", result.get("stderr", "")[-500:])
            return

        duration = result["duration_sec"]
        print(f"[TSS] Done in {duration:.1f}s, parsing levels...")

        # ---- Parse + axial mapping ----
        try:
            mapper = LevelMapper()
            levels = mapper.parse(out_dir, sag_nifti)
            levels = mapper.map_to_axial(levels, self.axial_slices)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.detect_levels_btn.setEnabled(True)
            self.detect_levels_btn.setText("Detect Levels")
            self.levels_status_label.setText(f"Parsing error: {e}")
            return

        # ---- Canal CSA per level (auto dural sac) ----
        try:
            from .totalspineseg.canal_csa import compute_all_level_csas
            canal_dir = Path(out_dir) / "step1_canal"
            # Find the single canal NIfTI without guessing basenames
            # (Path.stem doesn't handle .nii.gz double-extension correctly)
            candidates = list(canal_dir.glob("*.nii*"))
            if len(candidates) == 0:
                raise FileNotFoundError(f"No canal NIfTI in {canal_dir}")
            if len(candidates) > 1:
                # Exact stem match (startswith would wrongly match s201 for s2)
                in_base = Path(sag_nifti).name
                if in_base.endswith(".nii.gz"):
                    in_stem = in_base[:-7]
                elif in_base.endswith(".nii"):
                    in_stem = in_base[:-4]
                else:
                    in_stem = Path(sag_nifti).stem
                def _stem2(fname):
                    if fname.endswith(".nii.gz"):
                        return fname[:-7]
                    if fname.endswith(".nii"):
                        return fname[:-4]
                    return fname
                matched = [c for c in candidates if _stem2(c.name) == in_stem]
                canal_nifti = str(matched[0] if matched else candidates[0])
            else:
                canal_nifti = str(candidates[0])
            print(f"[CSA] Computing canal CSA from {canal_nifti}")
            levels = compute_all_level_csas(canal_nifti, levels, threshold=0.5)
            # Log results for debugging
            for lvl_name, lvl_info in levels.items():
                csa_d = lvl_info.get("canal_csa", {})
                if "csa_mm2" in csa_d:
                    print(f"[CSA] {lvl_name}: {csa_d['csa_mm2']:.1f} mm²")
                else:
                    print(f"[CSA] {lvl_name}: ERROR - {csa_d.get('error', '?')}")
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"[CSA] WARNING: canal CSA failed: {e}")
            # CSA hesaplanamasa bile devam - levels dict zaten elimizde

        # ---- Populate UI ----
        self.detected_levels = levels  # save for later analysis
        if hasattr(self, "analyze_all_btn"):
            self.analyze_all_btn.setEnabled(True)
        covered = sum(1 for v in levels.values() if v.get("axial_slice_idx") is not None)
        total = len(levels)
        self.levels_status_label.setText(
            f"Detected {total} levels - {covered} covered by axial volume "
            f"({duration:.0f}s)"
        )

        for name, info in levels.items():
            ax_idx = info.get("axial_slice_idx")
            wz = info["world_xyz"][2]
            # Canal CSA from auto dural sac segmentation (may be missing/erroneous)
            csa_info = info.get("canal_csa", {})
            if "csa_mm2" in csa_info:
                csa_str = f"  CSA: {csa_info['csa_mm2']:.0f} mm²"
            else:
                csa_str = ""
            if ax_idx is None:
                reason = info.get("out_of_range_reason", "gap")
                item_text = f"  {name:<10}  z={wz:6.1f} mm   OUT OF RANGE ({reason}){csa_str}"
                item = QListWidgetItem(item_text)
                item.setForeground(QColor(TXT_LT))
            else:
                item_text = (
                    f"  {name:<10}  z={wz:6.1f} mm   -> axial slice "
                    f"{ax_idx+1}/{len(self.axial_slices)}{csa_str}"
                )
                item = QListWidgetItem(item_text)
                item.setData(Qt.ItemDataRole.UserRole, ax_idx)  # for click-to-jump later
            self.levels_list.addItem(item)

        self.detect_levels_btn.setEnabled(True)
        self.detect_levels_btn.setText("Detect Levels")
        print(f"[TSS] Done. {covered}/{total} levels covered.")

        # Save canal NIfTI path for axial overlay - match by input basename
        try:
            canal_dir2 = Path(out_dir) / "step1_canal"
            cnd = list(canal_dir2.glob("*.nii*"))
            in_base = Path(sag_nifti).name
            if in_base.endswith(".nii.gz"):
                in_stem = in_base[:-7]
            elif in_base.endswith(".nii"):
                in_stem = in_base[:-4]
            else:
                in_stem = Path(sag_nifti).stem
            # Exact stem match - NOT startswith, because "sagittal_s201"
            # startswith "sagittal_s2" is True and would pick the wrong file.
            def _stem(fname):
                if fname.endswith(".nii.gz"):
                    return fname[:-7]
                if fname.endswith(".nii"):
                    return fname[:-4]
                return fname
            matched = [c for c in cnd if _stem(c.name) == in_stem]
            if matched:
                self.canal_nifti_path = str(matched[0])
            elif cnd:
                self.canal_nifti_path = str(cnd[0])
            else:
                self.canal_nifti_path = None

        except Exception:
            self.canal_nifti_path = None

        # ---- Overlay level lines on the sagittal display ----
        self._update_sagittal_level_overlay()

        # Refresh axial display so canal overlay appears on the current slice
        self._on_slider(self.current_slice_idx)

    def _update_sagittal_level_overlay(self):
        """Push detected_levels onto the sagittal display as horizontal lines."""
        if not getattr(self, "detected_levels", None):
            return
        if self.sagittal_data is None or self.sagittal_affine is None:
            return

        sag_aff = self.sagittal_affine
        si_axis = int(np.abs(sag_aff[2, :3]).argmax())
        si_step = sag_aff[2, si_axis]
        si_origin = sag_aff[2, 3]
        n_along = self.sagittal_data.shape[si_axis]

        # Stable per-level colors (RGB tuples). L3_body in distinct brown.
        LEVEL_COLORS = {
            "L1-L2":   (31, 119, 180),
            "L2-L3":   (44, 160,  44),
            "L3-L4":   (255, 127,  14),
            "L4-L5":   (214,  39,  40),
            "L5-S":    (148, 103, 189),
            "L3_body": (140,  86,  75),
        }

        level_lines = []
        for name, info in self.detected_levels.items():
            try:
                z_world = float(info["world_xyz"][2])
                pixel_idx = (z_world - si_origin) / si_step
                raw_fraction = pixel_idx / (n_along - 1)
                # Same inversion as _compute_sagittal_line_y for consistency.
                y_frac = 1.0 - raw_fraction
                y_frac = float(np.clip(y_frac, 0.0, 1.0))
            except Exception:
                continue
            covered = info.get("axial_slice_idx") is not None
            level_lines.append({
                "y_frac": y_frac,
                "color_rgb": LEVEL_COLORS.get(name, (255, 255, 0)),
                "label": name,
                "dashed": not covered,
            })

        # Re-render sagittal slice with overlays preserved.
        # Reuse current image; only update level_lines.
        if hasattr(self.sagittal_display, "_img") and self.sagittal_display._img is not None:
            self.sagittal_display._level_lines = level_lines
            self.sagittal_display._refresh()

    def _on_analyze_all_levels(self):
        """Run muscle + dural sac + stenosis analysis at every detected level."""
        if self.analyzer is None:
            return
        detected = getattr(self, "detected_levels", None)
        canal_path = getattr(self, "canal_nifti_path", None)
        if not detected:
            QMessageBox.warning(self, "Analyze All Levels",
                "Run Detect Levels first.")
            return
        if not canal_path:
            QMessageBox.warning(self, "Analyze All Levels",
                "No canal segmentation found. Run Detect Levels first.")
            return
        if not self.axial_slices:
            QMessageBox.warning(self, "Analyze All Levels",
                "Multi-level analysis requires a DICOM axial series.")
            return

        from .totalspineseg.multi_level_analyzer import MultiLevelAnalyzer

        demo = self._get_demographics()
        self.analyze_all_btn.setEnabled(False)
        self.analyze_all_btn.setText("Analyzing all levels...")
        self.status_label.setText("Analyzing all levels...")
        QApplication.processEvents()

        def _progress(msg):
            self.status_label.setText(msg)
            QApplication.processEvents()

        try:
            mla = MultiLevelAnalyzer(self.analyzer, self.axial_slices, canal_path)
            result = mla.analyze_all(
                detected,
                slice_nifti_producer=self._make_slice_nifti,
                demographics=demo,
                progress_callback=_progress,
            )
            self.multi_level_result = result
            self._show_multi_level_results(result)
            self.save_pdf_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)
            self.status_label.setText("Multi-level analysis complete.")
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Analyze All Levels",
                f"Multi-level analysis failed:\n{e}")
            self.status_label.setText("Multi-level analysis failed.")
        finally:
            self.analyze_all_btn.setEnabled(True)
            self.analyze_all_btn.setText("Analyze All Levels")

    def _show_multi_level_results(self, result):
        """Display multi-level results in the levels_list (temporary view)."""
        self.levels_list.clear()
        from PyQt6.QtWidgets import QListWidgetItem

        levels = result.get("levels", {})
        for lvl_name, data in levels.items():
            approx = " [approx]" if data.get("approx") else ""
            csa = data.get("canal_csa_mm2")
            sten = data.get("stenosis")
            n_musc = len(data.get("muscles", []))

            csa_str = f"{csa:.0f} mm2" if csa is not None else "N/A"
            line = f"{lvl_name}{approx}: dural sac {csa_str}, {n_musc} muscles"
            if sten and sten.get("flag"):
                line += f"  -> {sten['label']}"

            item = QListWidgetItem(line)
            # Color: red if stenosis flag, otherwise normal
            if sten and sten.get("flag"):
                item.setForeground(QColor(DANGER))
            self.levels_list.addItem(item)

        # Sarcopenia summary line
        sarc = result.get("sarcopenia")
        if sarc and sarc.get("result"):
            r = sarc["result"]
            pmi = r.get("pmi_cm2_per_m2")
            risk = r.get("risk_category", "Unknown")
            pmi_str = f"PMI {pmi}" if pmi else "PMI N/A"
            sline = f"L3 sarcopenia: {pmi_str}, risk {risk}"
            item = QListWidgetItem(sline)
            item.setForeground(QColor(NAVY))
            self.levels_list.addItem(item)

    def _make_slice_nifti(self, ax_idx):
        """Produce a single-slice NIfTI for the given axial slice index using
        the proven per-slice dcm2niix flow. Returns path str or None.
        Shared by single-slice Analyze and multi-level analysis."""
        if not self.axial_slices:
            return None
        if ax_idx < 0 or ax_idx >= len(self.axial_slices):
            return None
        slc = self.axial_slices[ax_idx]
        src_dicom = slc.get('source_path')
        if src_dicom is None or not Path(src_dicom).exists():
            return None
        import shutil as _shutil
        import subprocess as _sp
        from .dicom_loader import _resolve_dcm2niix
        tmp_dir = Path(tempfile.gettempdir()) / f"spinosarc_slice_{ax_idx}"
        if tmp_dir.exists():
            _shutil.rmtree(tmp_dir)
        tmp_dir.mkdir(parents=True)
        tmp_src_dir = tmp_dir / "src"
        tmp_src_dir.mkdir()
        _shutil.copy2(src_dicom, str(tmp_src_dir))
        dcm2niix_bin = _resolve_dcm2niix()
        _sp.run([dcm2niix_bin, '-o', str(tmp_dir), '-f', 'slice', str(tmp_src_dir)],
                capture_output=True, text=True, timeout=30)
        niftis = list(tmp_dir.glob('slice*.nii*'))
        return str(niftis[0]) if niftis else None

    def _on_analyze(self):
        if self.analyzer is None:
            return

        slice_path = None

        if self.axial_slices:
            # DICOM mode: use dcm2niix on the single source DICOM (proven reliable)
            slc = self.axial_slices[self.current_slice_idx]
            src_dicom = slc.get('source_path')
            if src_dicom is None or not Path(src_dicom).exists():
                QMessageBox.critical(self, "Analyze error",
                    "Source DICOM file not found for this slice.")
                return

            import shutil as _shutil
            tmp_dir = Path(tempfile.gettempdir()) / "spinosarc_one_slice"
            if tmp_dir.exists():
                _shutil.rmtree(tmp_dir)
            tmp_dir.mkdir(parents=True)
            tmp_src_dir = tmp_dir / "src"
            tmp_src_dir.mkdir()
            _shutil.copy2(src_dicom, str(tmp_src_dir))

            import subprocess as _sp
            from .dicom_loader import _resolve_dcm2niix
            dcm2niix_bin = _resolve_dcm2niix()
            r = _sp.run([dcm2niix_bin, '-o', str(tmp_dir), '-f', 'slice', str(tmp_src_dir)],
                       capture_output=True, text=True, timeout=30)
            niftis = list(tmp_dir.glob('slice*.nii*'))
            if not niftis:
                QMessageBox.critical(self, "Analyze error",
                    f"dcm2niix failed:\n{r.stdout[-500:]}\n{r.stderr[-500:]}")
                return
            slice_path = str(niftis[0])

        elif self.axial_data is not None:
            # NIfTI mode (eski yol)
            if self.is_volume:
                sl = self.axial_data[:, :, self.current_slice_idx:self.current_slice_idx + 1]
                tmp = Path(tempfile.gettempdir()) / "spinosarc_input.nii.gz"
                nib.Nifti1Image(sl, self.axial_affine, self.axial_header).to_filename(str(tmp))
                slice_path = str(tmp)
            else:
                slice_path = self.axial_path
        else:
            return

        demo = self._get_demographics()
        self.analyze_btn.setEnabled(False)
        self.status_label.setText("Analyzing...")
        QApplication.processEvents()
        try:
            result = self.analyzer.analyze(slice_path, demo)
            self._on_analysis_done(result)
        except Exception as e:
            import traceback; traceback.print_exc()
            self._on_analysis_error(str(e))

    def _on_analysis_done(self, result):
        self.last_result = result
        self.analyze_btn.setEnabled(True)
        self.save_pdf_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(True)
        self.status_label.setText("Analysis complete.")
        seg = result['segmentation_mask']
        if seg.ndim == 3:
            seg = seg[:, :, 0]

        # Compute auto canal CSA for the current slice (if Detect Levels was run).
        # This is the automatic dural sac segmentation - fast because it only
        # resamples the precomputed canal NIfTI (no TotalSpineSeg re-run).
        canal_mask_for_overlay = None
        if self.axial_slices:
            canal_mask_for_overlay = self._compute_canal_overlay_for_axial_slice(
                self.current_slice_idx
            )
            if canal_mask_for_overlay is not None:
                voxel_count = int(canal_mask_for_overlay.sum())
                slc = self.axial_slices[self.current_slice_idx]
                ps = slc.get('pixel_spacing', (1.0, 1.0))
                pixel_area = float(ps[0]) * float(ps[1])
                canal_csa = voxel_count * pixel_area
                self.csa_label.setText(
                    "Dural sac CSA: " + str(int(round(canal_csa))) + " mm2 (auto)"
                )

        # DICOM mode: overlay segmentation on original DICOM pixel array,
        # so pre- and post-analysis views match exactly.
        # Mask comes from dcm2niix LAS NIfTI -> rotate by k=1 to align with DICOM voxel.
        if self.axial_slices:
            dicom_img = self.axial_slices[self.current_slice_idx]['pixel_array']
            seg_aligned = np.rot90(seg, k=1)
            self.axial_display.set_image(dicom_img, overlay_mask=seg_aligned,
                                         rotation=self.axial_rotation,
                                         canal_overlay=canal_mask_for_overlay)
        else:
            # NIfTI mode (legacy)
            img = result['image_array']
            self.axial_display.set_image(img, overlay_mask=seg, rotation=self.axial_rotation)
        sarc = result['sarcopenia']
        risk = sarc['risk_category']
        cols = {'Low': SUCCESS, 'Moderate': WARNING, 'High': DANGER, 'Unknown': TXT_LT}
        col = cols.get(risk, TXT_LT)
        self.risk_label.setText(risk.upper() + " RISK")
        self.risk_label.setStyleSheet("font-size: 18px; font-weight: 700; color: " + col + ";")
        pmi = sarc.get('pmi_cm2_per_m2'); tpa = sarc['total_psoas_area_cm2']
        if pmi:
            self.pmi_label.setText("PMI: " + str(pmi) + " cm2/m2   TPA: " + str(tpa) + " cm2")
        else:
            self.pmi_label.setText("TPA: " + str(tpa) + " cm2 (demographics needed for PMI)")
        muscles = result['muscles']
        self.muscle_table.setRowCount(len(muscles))
        for i, m in enumerate(muscles):
            self.muscle_table.setItem(i, 0, QTableWidgetItem(m['name']))
            self.muscle_table.setItem(i, 1, QTableWidgetItem(str(int(m['csa_mm2']))))
            ff_item = QTableWidgetItem(str(round(m['fat_fraction'] * 100, 1)))
            ff = m['fat_fraction']
            ff_item.setForeground(QColor(SUCCESS if ff < 0.10 else (WARNING if ff < 0.25 else DANGER)))
            self.muscle_table.setItem(i, 2, ff_item)

    def _on_analysis_error(self, err):
        self.analyze_btn.setEnabled(True)
        self.status_label.setText("Analysis failed: " + str(err))
        QMessageBox.critical(self, "Analysis Error", str(err))

    def _gui_demographics_dict(self):
        """Collect demographics from the GUI inputs as a plain dict for reports."""
        age = self.age_input.value() or None
        sex = self.sex_input.currentText()
        sex = sex if sex in ("M", "F") else None
        h = self.height_input.value() or None
        w = self.weight_input.value() or None
        return {"age": age, "sex": sex, "height_cm": h, "weight_kg": w}

    def _on_save_pdf(self):
        # Multi-level report takes priority if a multi-level analysis exists.
        mlr = getattr(self, "multi_level_result", None)
        if mlr:
            pid = self.patient_id_input.text().strip() or "Patient"
            default_name = "SpinoSarc_MultiLevel_" + pid.replace(" ", "_") + ".pdf"
            out_path, _ = QFileDialog.getSaveFileName(
                self, "Save Multi-Level PDF Report", default_name,
                "PDF files (*.pdf)")
            if not out_path:
                return
            try:
                from .multi_level_report import export_multi_level_pdf
                self.status_label.setText("Generating multi-level PDF...")
                QApplication.processEvents()
                export_multi_level_pdf(
                    mlr, self._gui_demographics_dict(), pid, out_path)
                self.status_label.setText("PDF saved: " + out_path)
                QMessageBox.information(self, "Saved",
                    "Multi-level PDF saved to:\n" + out_path)
            except Exception as e:
                import traceback; traceback.print_exc()
                self.status_label.setText("PDF save failed: " + str(e))
                QMessageBox.critical(self, "Save error", str(e))
            return

        if not self.last_result:
            QMessageBox.warning(self, "No analysis", "Run analysis first.")
            return
        pid = self.patient_id_input.text().strip() or "Patient"
        default_name = "SpinoSarc_" + pid.replace(" ", "_") + ".pdf"
        out_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", default_name, "PDF files (*.pdf)")
        if not out_path:
            return
        try:
            from .single_page_report import generate_single_page_from_result
            self.status_label.setText("Generating PDF...")
            QApplication.processEvents()
            roi_pts = self.axial_display.get_roi_points_imgcoords()
            if roi_pts is not None:
                sp_x, sp_y = self._get_current_pixel_spacing()
                area_px = self._polygon_area_px(roi_pts)
                self.last_result['dural_sac_csa_mm2'] = round(area_px * sp_x * sp_y, 1)
            generate_single_page_from_result(pid, self.last_result, out_path)
            self.status_label.setText("PDF saved: " + out_path)
            QMessageBox.information(self, "Saved", "PDF saved to:\n" + out_path)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.status_label.setText("PDF save failed: " + str(e))
            QMessageBox.critical(self, "Save error", str(e))

    def _on_export_excel(self):
        """Metrikleri xlsx olarak dışa aktar."""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Missing dependency",
                "openpyxl is not installed. Install with: pip install openpyxl")
            return

        # Multi-level report takes priority if a multi-level analysis exists.
        mlr = getattr(self, "multi_level_result", None)
        if mlr:
            pid = self.patient_id_input.text().strip() or "Patient"
            default_name = "SpinoSarc_MultiLevel_" + pid.replace(" ", "_") + ".xlsx"
            out_path, _ = QFileDialog.getSaveFileName(
                self, "Export Multi-Level Metrics", default_name,
                "Excel files (*.xlsx)")
            if not out_path:
                return
            try:
                from .multi_level_report import export_multi_level_excel
                self.status_label.setText("Generating multi-level Excel...")
                QApplication.processEvents()
                export_multi_level_excel(
                    mlr, self._gui_demographics_dict(), pid, out_path)
                self.status_label.setText("Excel saved: " + out_path)
                QMessageBox.information(self, "Saved",
                    "Multi-level Excel saved to:\n" + out_path)
            except Exception as e:
                import traceback; traceback.print_exc()
                self.status_label.setText("Excel save failed: " + str(e))
                QMessageBox.critical(self, "Export error", str(e))
            return

        if not self.last_result:
            QMessageBox.warning(self, "No analysis", "Run analysis first.")
            return

        pid = self.patient_id_input.text().strip() or "Patient"
        default_name = "SpinoSarc_" + pid.replace(" ", "_") + ".xlsx"
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Export Metrics to Excel", default_name, "Excel files (*.xlsx)")
        if not out_path:
            return

        try:
            self.status_label.setText("Generating Excel...")
            QApplication.processEvents()
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime

            # ROI'yi guncelle (PDF ile ayni mantik)
            roi_pts = self.axial_display.get_roi_points_imgcoords()
            if roi_pts is not None:
                sp_x, sp_y = self._get_current_pixel_spacing()
                area_px = self._polygon_area_px(roi_pts)
                self.last_result['dural_sac_csa_mm2'] = round(area_px * sp_x * sp_y, 1)

            wb = Workbook()
            ws = wb.active
            ws.title = "SpinoSarc Report"

            # Stiller
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill("solid", fgColor="0A4F8C")
            section_font = Font(bold=True, size=11, color="0A4F8C")
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Header
            ws['A1'] = "SpinoSarc Quantitative Analysis Report"
            ws['A1'].font = Font(bold=True, size=14, color="0A4F8C")
            ws.merge_cells('A1:D1')
            ws['A2'] = "Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M")
            ws['A2'].font = Font(italic=True, size=10, color="666666")
            ws.merge_cells('A2:D2')

            row = 4
            # Patient demographics
            ws.cell(row=row, column=1, value="Patient Demographics").font = section_font
            row += 1
            demo_rows = [
                ("Patient ID", pid),
                ("Age", str(self.age_input.value()) + " yrs" if self.age_input.value() > 0 else "-"),
                ("Sex", self.sex_input.currentText() if self.sex_input.currentText() != "-" else "-"),
                ("Height", str(self.height_input.value()) + " cm" if self.height_input.value() > 0 else "-"),
                ("Weight", str(self.weight_input.value()) + " kg" if self.weight_input.value() > 0 else "-"),
            ]
            for label, val in demo_rows:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=val)
                row += 1

            row += 1
            # Sarcopenia summary
            sarco = self.last_result.get('sarcopenia', {}) or {}
            ws.cell(row=row, column=1, value="Sarcopenia Assessment").font = section_font
            row += 1
            sarco_rows = [
                ("Total Psoas Area (TPA)", f"{sarco.get('total_psoas_area_cm2', 0):.2f} cm2"),
                ("Psoas Muscle Index (PMI)",
                 f"{sarco.get('pmi_cm2_per_m2', 0):.2f} cm2/m2" if sarco.get('pmi_cm2_per_m2') else "demographics needed"),
                ("Risk Category", sarco.get('risk_category', 'Unknown')),
            ]
            for label, val in sarco_rows:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=val)
                row += 1

            # Dural sac CSA (eger varsa)
            dural = self.last_result.get('dural_sac_csa_mm2')
            if dural is not None:
                row += 1
                ws.cell(row=row, column=1, value="Dural Sac CSA").font = section_font
                row += 1
                ws.cell(row=row, column=1, value="CSA").font = Font(bold=True)
                ws.cell(row=row, column=2, value=f"{dural:.1f} mm2")
                row += 1
                if dural < 75:
                    stenosis = "Severe stenosis (<75 mm2)"
                elif dural < 100:
                    stenosis = "Relative stenosis (<100 mm2)"
                else:
                    stenosis = "Normal (>=100 mm2)"
                ws.cell(row=row, column=1, value="Classification").font = Font(bold=True)
                ws.cell(row=row, column=2, value=stenosis)
                row += 1

            row += 1
            # Per-muscle tablo
            ws.cell(row=row, column=1, value="Muscle Measurements").font = section_font
            row += 1
            headers = ["Muscle", "CSA (mm2)", "FF (%)"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
                c.border = border
            row += 1
            for m in self.last_result.get('muscles', []):
                ws.cell(row=row, column=1, value=m.get('name', '?')).border = border
                csa = m.get('csa_mm2', 0)
                ws.cell(row=row, column=2, value=round(csa, 1)).border = border
                ff = m.get('fat_fraction', None)
                if ff is not None:
                    ws.cell(row=row, column=3, value=round(ff, 1)).border = border
                else:
                    ws.cell(row=row, column=3, value="-").border = border
                row += 1

            # Sutun genislikleri
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 16

            wb.save(out_path)
            self.status_label.setText("Excel saved: " + out_path)
            QMessageBox.information(self, "Saved", "Excel saved to:\n" + out_path)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.status_label.setText("Excel save failed: " + str(e))
            QMessageBox.critical(self, "Export error", str(e))

    def _on_new_case(self):
        """Yeni vaka icin tum state'i temizle, baslangic durumuna don."""
        # State temizligi
        self.axial_slices = []
        self.axial_data = None
        self.axial_affine = None
        self.axial_header = None
        self.axial_path = None
        self.sagittal_data = None
        self.sagittal_affine = None
        self.sagittal_mid_idx = None
        self.current_sag_idx = 0
        self.is_volume = False
        self.current_slice_idx = 0
        self.last_result = None
        # UI temizligi
        self.axial_display.set_image(None)
        self.sagittal_display.set_image(None)
        self.axial_display.clear_roi()
        self.viewer_widget.hide()
        self.drop_zone.show()
        self.file_info.setText("")
        # Demografi reset
        self.patient_id_input.clear()
        self.age_input.setValue(0)
        self.sex_input.setCurrentIndex(0)
        self.height_input.setValue(0)
        self.weight_input.setValue(0)
        # Sonuc tablosu temizle
        self.muscle_table.setRowCount(0)
        self._reset_csa_display()
        # Butonlar
        self.analyze_btn.setEnabled(False)
        self.save_pdf_btn.setEnabled(False)
        self.export_excel_btn.setEnabled(False)
        self.new_case_btn.setEnabled(False)
        self.status_label.setText("Ready - drop NIfTI file(s) or DICOM folder to begin")


def main():
    app = QApplication(sys.argv)
    win = SpinoSarcWindow()
    win.show(); win.raise_(); win.activateWindow()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
